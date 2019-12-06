#IMPORTACIONES NECESARIAS, EN CASO DE QUE SE MANDE UN ERROR DEBIDO A QUE NO SE ENCUENTRA UN MÓDULO INSTALARLAS CON PIP
import os
from docx import Document
from openpyxl import load_workbook
from openpyxl import Workbook

pathBase = "C:/Users/rdominguezp/Desktop/"#PATH DONDE SE VA A GUARDAR LA BASE DE DATOS (no olvidar las /)
pathCuestionarios = "C:/Users/rdominguezp/Desktop/python/Cuestionarios/AutomatizadoTotal/"#PATH DE DONDE VAMOS A LEER LOS CUESTIONARIOS DE WORD (no olvidar las /)
nombreDeLaBase = "BASE_DE_DATOS.xlsx"#NOMBRE QUE LE VAMOS A DAR A LA BASE DE DATOS (no olvidar el .xlsx)



#==================================== INICIO DE LA FUNCIÓN EscribirTitulos ====================================

def EscribirTitulos(ws,titulo,ren):

    if ren == 0 :#SOLO SI SE ESCRIBE EN EL PRIMER RENGLÓN DE LA BASESE ESCRIBIRÁN LOS ENCABEZADOS
        ws.cell(row = ren+1 , column = 1 , value = "ENT")
        ws.cell(row = ren+1 , column = 3 , value = "FECHA")
        ws.cell(row = ren+1 , column = 2 , value = "GEN")

    #SE ESCRIBIRÁ EL NOMBRE DEL ARCHIVO DE DONDE SE ESTÁ LEYENDO EL CUESTIONARIO (título)
    ws.cell(row = ren+2 , column = 1 , value = titulo)
    ws.cell(row = ren+3 , column = 1 , value = titulo)
    ws.cell(row = ren+4 , column = 1 , value = titulo)
    ws.cell(row = ren+5 , column = 1 , value = titulo)
            
    ws.cell(row = ren+2 , column = 2 , value = "M")
    ws.cell(row = ren+3 , column = 2 , value = "H")
    ws.cell(row = ren+4 , column = 2 , value = "M")
    ws.cell(row = ren+5 , column = 2 , value = "H")

    
    ws.cell(row = ren+2 , column = 3 , value = "2018")
    ws.cell(row = ren+3 , column = 3 , value = "2018")
    ws.cell(row = ren+4 , column = 3 , value = "2019")
    ws.cell(row = ren+5 , column = 3 , value = "2019")    

#====================================== FIN DE LA FUNCIÓN EscribirTitulos ======================================



#================================= INICIO DE LA FUNCIÓN ExtraccionCuestionarios =================================
def ExtraccionCuestionarios(nombreArchivo,ws,ren):    
    
    #SE INICIALIZA LA VARIABLE CABECERA. SE USARÁ PARA SEPARAR LAS PREGUNTAS DE LOS TITULOS
    cabeceras = []

    #SE CREA UNA LISTA FINAL DONDE SE GUARDARÁ CADA PREGUNTA CON SUS RESPECTIVOS RESULTADOS
    listaF = []


    #============================SE ABRE EL DOCUMENTO INDICADO (Word del cuestionario)============================
    wordDoc = Document(pathCuestionarios + nombreArchivo + '.docx')

    #====================SE ITERA SOBRE CADA TABLA, CADA RENGLÓN Y CADA CELDA DEL RENGLÓN====================
    for table in wordDoc.tables:
        for row in table.rows:            
            indice = 0 #SE CREA UN INDICE PARA SABER EN QUE CELDA ESTAMOS (LAS PREGUNTAS SIEMPRE OCUPAN DOS CELDAS)
            rPreg = ""
            rRes = []        
            for cell in row.cells:
                #SI LA CELDA UNO DEL RENGLÓN ESTA VACÍA O SON SOLO MAYÚSCULAS SE TOMARÁ COMO CABECERA
                if row.cells[0].text.isupper() or row.cells[0].text == "":                
                    cabeceras.append(cell.text)

                else:
                    if indice < 2 :
                        #SI EL TEXTO SE ENCUENTRA DENTRO DE LAS PRIMERAS DOS COLUMNAS SE TOMARÁ COMO PREGUNTA
                        rPreg = row.cells[0].text
                        indice = indice + 1

                    else:
                        #EN CASO CONTRARIO SE AGREGA EL TEXTO A LA LISTA DE RESPUESTAS
                        rRes.append(cell.text)                    

            #AL FINAL DE CADA RENGLÓN SE AGREGARÁN A LA LISTA FINAL TANTO LA PREGUNTA COMO SUS RESPECTIVAS RESPUESTAS
            if rPreg != "":
                listaF.append({"Pregunta":rPreg,"Respuestas":rRes})
    
    i=0 #INDICE

    for l in listaF:
        #POR CADA ELEMENTO EN LA LISTA FINAL, SE ESCRIBE LA PREGUNTA EN LA CELDA ADECUADA
        #ESTO SE HARÁ POR CADA CUESTIONARIO, SIN EMBARGO SE SOBREESCRIBIRA EL PRIMER RENGLÓN CADA VEZ
        ws.cell(row = 1 , column = i+4 , value = l["Pregunta"])
        
        #SE ITERA SOBRE CADA RESPUESTA QUE CONTENGA LA LISTA SE ESCRIBIRÁ DE ACUERDO AL RENGLÓN QUE SE HAYA OBTENIDO COMO PARÁMETRO
        for j in range(len(l["Respuestas"])):
            ws.cell(row = ren+j+2 , column = i+4 , value = l ["Respuestas"][j])
            

        i = i+1        
        

    EscribirTitulos(ws,nombreArchivo,ren)#FUNC


    print("Se editó el Excel con la información de: "+nombreArchivo)    

#================================== FIN DE LA FUNCIÓN ExtraccionCuestionarios ==================================



#HACE UNA LISTA CON TODOS LOS ARCHIVOS QUE SE ENCUENTRAN EN LA CARPETA PARA LOS CUSTIONARIOS Y PARA LA BASE
archivosDirectorioBase = os.listdir(pathBase)
archivosDirectorioCuestionarios = os.listdir(pathCuestionarios)


if nombreDeLaBase in archivosDirectorioBase:#SE BUSCA SI LA BASE ESTA CREADA O NO

    print("Entra a la carga de un archivo creado")

    wb = load_workbook(filename = pathBase+nombreDeLaBase) #SE CARGA LA BASE EN CASO DE EXISTIR
    ws = wb.active#SE OBTIENE LA PRIMERA HOJA

    ren = ws.max_row - 1

    for archivo in archivosDirectorioCuestionarios:
        
        if archivo[-5:] == ".docx": #SE VERIFICA SI ES UN ARCHIVO DE WORD
            
            nombreArchivo = archivo[:-5]#SE OBTIENE EL NOMBRE DEL ARCHIVO SIN LA EXTENSIÓN
            
            #SE LLAMA LA FUNCIÓN MANDANDO COMO PARÁMETRO EL NOMBRE DEL ARCHIVO, LA HOJA QUE ESTAMOS EDITANDO EN EL EXCEL Y EL RENGLÓN EN QUE SE DEBE EMPEZAR A ESCRIBIR
            ExtraccionCuestionarios(nombreArchivo,ws,ren)
            ren = ren + 4 #CADA QUE TERMINE DE ESCRIBIR UN CUESTIONARIO SE AUNMENTARÁ EN 4 EL RENGLÓN (debido al formato especificado)


    """ SE GUARDA LA BASE DE DATOS (path + el nombre) ESTE ESTE PASO ES PROBABLEMENTE EL MÁS IMPORTANTE, NO SE
    GUARDARÁ NINGÚN CAMBIO AUNQUE EL PROGRAMA HAYA MOSTRADO QUE SE HAN CREADO ARCHIVOS DE EXCEL """
    wb.save(pathBase+nombreDeLaBase)
    
else:
    #Método similar a lo anterior en caso de que no exista el archivo de la base de datos
    print("No se encontró el archivo especificado")
    wb = Workbook()#SE CREA UN NUEVO EXCEL
    ws = wb.active

    ren = 0

    for archivo in archivosDirectorioCuestionarios:
        #==============================SE BUSCA POR UN ARCHIVO DE WORD==============================
        if archivo[-5:] == ".docx":
            #SE OBTIENE EL NOMBRE DEL ARCHIVO SIN LA EXTENSIÓN
            nombreArchivo = archivo[:-5]
            #SE LLAMA LA FUNCIÓN MANDANDO COMO PARÁMETRO EL NOMBRE DEL ARCHIVO
            ExtraccionCuestionarios(nombreArchivo,ws,ren)
            ren = ren + 4


    """ SE GUARDA LA BASE DE DATOS (path + el nombre) ESTE ESTE PASO ES PROBABLEMENTE EL MÁS IMPORTANTE, NO SE
    GUARDARÁ NINGÚN CAMBIO AUNQUE EL PROGRAMA HAYA MOSTRADO QUE SE HAN CREADO ARCHIVOS DE EXCEL """
    wb.save(pathBase+nombreDeLaBase)
    print("Nuevo Archivo Creado")


print("\n\n\n")
print("SE HAN CREADO TODOS LOS ARCHIVOS DE EXCEL")